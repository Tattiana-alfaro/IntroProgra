import xlrd 
from xlsxwriter import Workbook 
import xlwt
import openpyxl

path = "TestData//libros.xlsx"
wb = xlrd.open_workbook(path)
sheet = wb.sheet_by_index(0)

wb2 = Workbook(path)
sheet2 = wb2.add_worksheet()

miLista = []


class Libro:
    idLibro = ""
    nombre = ""
    genero = ""
    autor = ""
    disponibilidad = ""
    
    def __init__(self, idLibro, nombre, genero, autor,disponibilidad):
        self.idLibro = idLibro
        self.nombre = nombre
        self.genero = genero
        self.autor = autor
        self.disponibilidad= disponibilidad

# Funcion para cargar lista de libros
def cargarListaLibros():
    miLista.clear()
    for i in range(sheet.nrows):
        libro = Libro(sheet.cell_value(i, 0), sheet.cell_value(i, 1), sheet.cell_value(i, 2), sheet.cell_value(i, 3),sheet.cell_value(i, 4))
        miLista.append(libro)
    return miLista


# Funcion para imprimir la lista de libros
def imprirListaLibro():
    for Libro in cargarListaLibros():
        print(Libro.idLibro, " - ", Libro.nombre, " - ", Libro.genero, " - ", Libro.autor, Libro.disponibilidad)

#Funcion para buscar libros
def buscarLibro():

    encontrado = False
    texto = """
    ****************************************
    Por cual atributo le gustaria buscar
    ****************************************
    1. Codigo Libro
    2. Nombre
    3. Genero
    4. Autor
    ****************************************
    """
    print(texto)
    opcion = input("Digite una opcion: ")
    miLista = cargarListaLibros()
    if (int(opcion) == 1):
        attributo = input("Digite el codigo que desea buscar: ")
        for Libro in miLista:
            if str(attributo.capitalize()) in str(Libro.idLibro).capitalize():
                encontrado = True
                print("Se ha encontrado el siguiente registro: ")
                print("CODIGO: ", Libro.idLibro) 
                print("NOMBRE: ", Libro.nombre) 
                print("GENERO: ", Libro.genero) 
                print("AUTOR: ", Libro.autor) 
        
    if (int(opcion) == 2):
        attributo = input("Digite el nombre que desea buscar: ")
        for Libro in miLista:
            if str(attributo.capitalize()) in str(Libro.nombre).capitalize():
                encontrado = True
                print("Se ha encontrado el siguiente registro: ")
                print("CODIGO: ", Libro.idLibro) 
                print("NOMBRE: ", Libro.nombre) 
                print("GENERO: ", Libro.genero) 
                print("AUTOR: ", Libro.autor) 
        

    if (int(opcion) == 3):
        attributo = input("Digite el genero que desea buscar: ")
        for Libro in miLista:
            if str(attributo.capitalize()) in str(Libro.genero).capitalize():
                encontrado = True
                print("Se ha encontrado el siguiente registro: ")
                print("CODIGO: ", Libro.idLibro) 
                print("NOMBRE: ", Libro.nombre) 
                print("GENERO: ", Libro.genero) 
                print("AUTOR: ", Libro.autor) 

    if (int(opcion) == 4):
        attributo = input("Digite el autor que desea buscar: ")
        for Libro in miLista:
            if str(attributo.capitalize()) in str(Libro.autor).capitalize():
                encontrado = True
                print("Se ha encontrado el siguiente registro: ")
                print("CODIGO: ", Libro.idLibro) 
                print("NOMBRE: ", Libro.nombre) 
                print("GENERO: ", Libro.genero) 
                print("AUTOR: ", Libro.autor) 
        
    else:
        if encontrado == False:
            print("No se ha encontrado ninguna coincidencia")
    miLista.clear()
    
# Funcion para mostrar disponibilidad de los libros
def cargardisponibilidad():
    miLista = cargarListaLibros()
    for i in range(1,sheet.nrows):

        print("id", miLista[i].idLibro," ",miLista[i].nombre,":",miLista[i].disponibilidad)
        print(miLista[i].nombre,":",miLista[i].disponibilidad)
        

# Funcion para solicitar libro
def solicitarlibro(id_by_user):
 
    miLista = cargarListaLibros()
    
    for i in range(1, sheet.nrows ):

        if (int(miLista[i].idLibro) == int(id_by_user ) ) :

            if (miLista[i].disponibilidad == "si"):
                print("libro solicitado exitosamente")
                file = 'TestData/libros.xlsx'
                wb2 = openpyxl.load_workbook(filename=file)
                # Seleciono la Hoja
                ws = wb2['Sheet1']
                # Posicion
                lugar="E"+str(i+1)
                ws[lugar] = "no" #modifica el excel
                miLista[i].disponibilidad="no" #modifica la lista de python
                # Escribirmos en el Fichero
                wb2.save(file)
   
            else:
                print("Lo sentimos,el libro no está disponible")

                